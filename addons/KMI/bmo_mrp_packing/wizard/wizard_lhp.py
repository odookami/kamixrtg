# -*- coding: utf-8 -*-

import io
import os
import xlwt
import base64
import pandas as pd
from openpyxl import load_workbook
from datetime import date, datetime
from odoo.exceptions import ValidationError
from dateutil.relativedelta import relativedelta
from datetime import date, datetime, timedelta
from odoo import api, fields, models, SUPERUSER_ID

class MrpPackingReport(models.TransientModel):
    _name = 'mrp.packing.report'
    _description = 'Laporan Hasil Packing'

    tipe = fields.Selection([('all', 'ALL'), ('dep', 'Department')], 
        string='Tipe Report', default='all', required=True)
    user_id = fields.Many2one('res.users','Created By', required=True, \
        readonly=True, default=lambda self: self.env.uid)
    department_id = fields.Many2one('hr.department', 'Departement')
    state = fields.Selection([('all', 'All Status'), ('draft', 'Draft'), 
        ('verify', 'Waiting'), ('done', 'Done'), ('cancel', 'Rejected')], 
        string='Status', default='draft', required=True)
    date_from = fields.Date('Date From', required=True)
    date_to = fields.Date('Date To', required=True)
    data_file = fields.Binary('File')
    name = fields.Char('File Name')

    def _get_data_bhp(self, domain):
        bhp_obj = self.env['mrp.production.packing']
        bhp_ids = bhp_obj.search(domain)
        return bhp_ids


    def action_export_xls(self):
        PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
        BASE_DIR = os.path.dirname(PROJECT_ROOT)
        PATH_DIR = '/static/src/doc/'
        FILE_DIR = BASE_DIR + PATH_DIR
        FILE_NAME = 'LHP.xlsx'
        wb = load_workbook(FILE_DIR + FILE_NAME)
        sheet = wb.active
        # book = xlwt.Workbook()
        # sheet = book.add_sheet("Sheet 1", cell_overwrite_ok=True)
        now = datetime.now()

        style_no_bottom_border = xlwt.easyxf('font: name Calibri, height 260, \
            bold 1;align: horz center')

        header_style = xlwt.easyxf('font:height 200;pattern: pattern solid, fore_color gray25;'
                                   'align: horiz center;font: color black; font:bold True;'
                                   "borders: top thin,left thin,right thin,bottom thin")

        left_header_style = xlwt.easyxf('font:height 200;pattern: pattern solid, fore_color gray25;'
                                        'align: horiz left;font: color black; font:bold True;'
                                        "borders: top thin,left thin,right thin,bottom thin")
        text_left = xlwt.easyxf('font:height 200; align: horiz left;')
        text_center = xlwt.easyxf(
            'font: name Calibri; align: vert centre, horz center')
        style_bold_left = xlwt.easyxf('font: name Calibri, height 240, bold 1; \
            align: horz left')
        style_header2 = xlwt.easyxf('font: name Calibri, height 210, bold 1; \
            pattern: pattern solid, fore_colour yellow;borders: left thin, \
            right thin, top thin, bottom thin;align: vert centre, horz center')
        style_no_bold = xlwt.easyxf(
            'font: name Calibri; align: vert centre, horz center')
        style_no_bold.num_format_str = '#,##0.0000'

        domain = [('date', '>=', self.date_from), ('date', '<=', self.date_to),
                  ('state', '!=', 'draft')]
        bhp_line_ids = self._get_data_bhp(domain).mapped('mrp_packing_line').filtered(
            lambda l: l.mrp_packing_id != False)
        sortir_obj = self.env['mrp.sortir.report']
        data_dict = {}
        for fg in bhp_line_ids.filtered(lambda l: l.packing_type == 'Banded'):
            move_line_do = self.env['stock.move.line'].search([
                ('package_id.name', '=', fg.package_id.name), ('lot_id', '=', fg.lot_producing_id.id), ('state', '=', 'done')]).filtered(
                lambda l: l.location_id.usage == 'customer' or l.location_dest_id.usage == 'customer')
            if fg.okp_id not in data_dict:
                data_dict[fg.okp_id] = {fg.package_id.name: {
                    'Filling': [],
                    'Banded': [fg],
                    'DO': []
                }}
                for r in move_line_do:
                    data_dict[fg.okp_id][fg.package_id.name]['DO'].append(r)
            else:
                if fg.package_id.name not in data_dict[fg.okp_id]:
                    data_dict[fg.okp_id][fg.package_id.name] = {
                        'Filling': [],
                        'Banded': [fg],
                        'DO': []
                    }
                    for r in move_line_do:
                        data_dict[fg.okp_id][fg.package_id.name]['DO'].append(r)
                else:
                    data_dict[fg.okp_id][fg.package_id.name]['Banded'].append(fg)
                    for r in move_line_do:
                        data_dict[fg.okp_id][fg.package_id.name]['DO'].append(r)
        
        for wip in bhp_line_ids.filtered(lambda l: l.packing_type == 'Filling'):
            if wip.okp_id in data_dict:
                move_line_do = self.env['stock.move.line'].search([
                    ('package_id.name', '=', wip.package_id.name), ('lot_id', '=', wip.lot_producing_id.id), ('state', '=', 'done')]).filtered(
                    lambda l: l.location_id.usage == 'customer' or l.location_dest_id.usage == 'customer')
                if wip.package_id.name not in data_dict[wip.okp_id]:
                    data_dict[wip.okp_id][wip.package_id.name] = {
                        'Filling': [wip],
                        'Banded': [],
                        'DO': []
                    }
                    for r in move_line_do:
                        data_dict[wip.okp_id][wip.package_id.name]['DO'].append(r)
                else:
                    data_dict[wip.okp_id][wip.package_id.name]['Filling'].append(wip)
                    for r in move_line_do:
                        data_dict[wip.okp_id][wip.package_id.name]['DO'].append(r)
        
        num = 8
        for k, v in data_dict.items():
            total_output = sampling_ex_qty = total_sortir_1 = total_sortir_2 = 0
            t = -1
            sort_dict = dict(sorted(v.items(), key=lambda x: x[0]))
            for u, i in sort_dict.items():
                num += 1
                t += 1
                nilai_a = f'A{num}'
                nilai_b = f'B{num}'
                nilai_c = f'C{num}'
                nilai_d = f'D{num}'
                nilai_e = f'E{num}'
                # nilai_f = f'F{num}'
                nilai_g = f'G{num}'
                nilai_h = f'H{num}'
                nilai_i = f'I{num}'
                nilai_j = f'J{num}'
                nilai_k = f'K{num}'

                nilai_o = f'O{num}'
                nilai_p = f'P{num}'
                nilai_q = f'Q{num}'
                nilai_r = f'R{num}'
                nilai_s = f'S{num}'
                nilai_t = f'T{num}'
                nilai_u = f'U{num}'
                nilai_v = f'V{num}'
                nilai_w = f'W{num}'
                nilai_x = f'X{num}'
                nilai_y = f'Y{num}'
                nilai_z = f'Z{num}'

                nilai_aa = f'AA{num}'
                nilai_ab = f'AB{num}'
                nilai_ac = f'AC{num}'
                nilai_ad = f'AD{num}'
                nilai_ae = f'AE{num}'
                nilai_af = f'AF{num}'
                nilai_ag = f'AG{num}'
                nilai_ah = f'AH{num}'
                nilai_ai = f'AI{num}'
                nilai_aj = f'AJ{num}'
                nilai_ak = f'AK{num}'
                nilai_al = f'AL{num}'
                nilai_am = f'AM{num}'
                nilai_an = f'AN{num}'
                nilai_ao = f'AO{num}'
                nilai_ap = f'AP{num}'
                nilai_aq = f'AQ{num}'
                nilai_ar = f'AR{num}'
                nilai_as = f'AS{num}'
                nilai_at = f'AT{num}'

                nilai_bb = f'BB{num}'
                nilai_bc = f'BC{num}'
                nilai_bd = f'BD{num}'
                nilai_be = f'BE{num}'

                nilai_bg = f'BG{num}'
                nilai_bh = f'BH{num}'
                nilai_bi = f'BI{num}'
                nilai_bj = f'BJ{num}'
                nilai_bk = f'BK{num}'
                nilai_bl = f'BL{num}'
                nilai_bm = f'BM{num}'
                sheet[nilai_e] = u
                for l in i['Banded']:
                    bmb = l.mrp_packing_id._get_batch_mrp(l.okp_id, 'Banded')
                    sheet[nilai_a] = l.product_id.code
                    sheet[nilai_b] = l.product_id.name
                    sheet[nilai_s] = bmb.number_bo or ''
                    sheet[nilai_t] = round(l.product_id.expiration_time / 30, 2)
                    sheet[nilai_u] = bmb.bo_fee or ''

                    sum_total_output_banded = sum(
                        x.total_output_banded for x in l if x.mrp_packing_id.okp_id.id == l.mrp_packing_id.okp_id.id and x.mrp_packing_id.packing_type == l.mrp_packing_id.packing_type)
                    sheet[nilai_bb] = l.qty_banded_in_ct
                    sheet[nilai_bc] = l.qty_banded_in_pcs
                    sheet[nilai_bd] = l.total_output_banded
                    sheet[nilai_be] = sum_total_output_banded
                    sheet[nilai_bg] = l.picking_id.date_done or ''
                for l in i['Filling']:
                    bmr = l.mrp_packing_id._get_batch_mrp(l.okp_id, 'Filling')
                    sheet[nilai_c] = bmr.name
                    sheet[nilai_d] = l.lot_producing_id.name
                    sheet[nilai_g] = l.start
                    sheet[nilai_h] = l.finish
                    sheet[nilai_i] = l.first_check_code or ''
                    sheet[nilai_j] = l.first_check_time or ''
                    sheet[nilai_k] = l.first_check_count or ''
                    sheet[nilai_p] = str(l.first_check_code)[-1]
                    sheet[nilai_q] = k.name
                    sheet[nilai_r] = bmr.number_bo or ''
                    sampling_ex_qty = bmr.sampling_ex_qty

                    bmm = l.mrp_packing_id._get_batch_mrp(l.okp_id, 'Mixing')
                    sheet[nilai_v] = bmm.planned_start or ''
                    sheet[nilai_w] = bmm.actual_complete_date + timedelta(hours=7) if bmm.actual_complete_date else ''
                    sheet[nilai_x] = bmm.goods_qty or 0

                    sheet[nilai_y] = bmr.planned_start or ''
                    sheet[nilai_z] = bmr.actual_complete_date + timedelta(hours=7) if bmr.actual_complete_date else ''
                    sheet[nilai_aa] = l.lot_producing_id.name or ''
                    sheet[nilai_ab] = l.lot_producing_id.expiration_date or ''
                    sheet[nilai_ac] = l.qty_in_ct or 0
                    sheet[nilai_ad] = l.qty_in_pcs or 0
                    sheet[nilai_ae] = l.total_output or 0
                    total_output += l.total_output

                    sheet[nilai_ag] = l.location_dest_id.name or ''

                    sortir_id = sortir_obj.search(
                        [('mrp_packing_id', '=', l.mrp_packing_id.id), ('package_id.name', '=', u),('state','=','done')])
                    mix_date = bmr.actual_complete_date + timedelta(hours=7) if bmr.actual_complete_date else  ''
                    plan_date = ' ' if not mix_date else str(mix_date + relativedelta(days=4))
                    for x in sortir_id:
                        if x.sort_seq == '1':
                            sheet[nilai_ai] = plan_date
                            sheet[nilai_aj] = x.actual_date or ' '
                            sheet[nilai_ak] = x.output_cb or 0
                            sheet[nilai_al] = x.output_pcs or 0
                            sheet[nilai_am] = x.output_total or 0
                            total_sortir_1 += x.output_total
                        else:
                            sheet[nilai_ao] = plan_date
                            sheet[nilai_ap] = x.actual_date or ' '
                            sheet[nilai_aq] = x.output_cb or 0
                            sheet[nilai_ar] = x.output_pcs or 0
                            sheet[nilai_as] = x.output_total or 0
                            total_sortir_2 += x.output_total

                for d in i['DO']:
                    if len(d) > 1:
                        for t in d:
                            sheet[nilai_bh] = t.picking_id.sale_id.client_order_ref
                            sheet[nilai_bi] = t.qty_done
                            sheet[nilai_bj] = "Sesuai" if t.qty_done > 0 else ''
                            sheet[nilai_bk] = t.picking_id.date_done
                            sheet[nilai_bl] = t.picking_id.name
                            sheet[nilai_bm] = t.qty_done
                            num += 1
                    else:
                        move_id = d.move_id
                        sheet[nilai_bh] = d.picking_id.sale_id.client_order_ref
                        sheet[nilai_bi] = d.qty_done
                        sheet[nilai_bj] = "Sesuai" if d.qty_done > 0 else ''
                        sheet[nilai_bk] = d.picking_id.date_done
                        sheet[nilai_bl] = d.picking_id.name
                        sheet[nilai_bm] = d.qty_done
            sheet.merge_cells(f'AF{num-t}:AF{num}')
            sheet.merge_cells(f'AH{num-t}:AH{num}')
            sheet.merge_cells(f'AN{num-t}:AN{num}')
            sheet.merge_cells(f'AN{num-t}:AN{num}')
            top_left_cell = sheet[f'AF{num-t}']
            top_left_cell.value = total_output
            top_left_cell = sheet[f'AH{num-t}']
            top_left_cell.value = sampling_ex_qty
            top_left_cell = sheet[f'AN{num-t}']
            top_left_cell.value = total_sortir_1
            sheet.merge_cells(f'AT{num-t}:AT{num}')
            top_left_cell = sheet[f'AT{num-t}']
            top_left_cell.value = total_sortir_2
        sheet['CB2'] = f': {now.strftime("%Y-%m-%d")}'
        setting = self.env['res.company']._company_default_get('res.config.settings')
        sheet['BN2'] = f': {setting.date_report_lhp.strftime("%d-%m-%Y")}'
        file_data = io.BytesIO()
        wb.save(file_data)
        out = base64.encodebytes(file_data.getvalue())
        filename = 'LHP - ' + str(self.date_to.strftime('%B')) + \
            ' - ' + str(self.date_to.strftime('%Y')) + '.xlsx'
        self.write({'data_file': out, 'name': filename})

        model = self._name
        field_file = 'data_file'
        content = 'web/content/?model=%s&field=%s' % (model, field_file)
        download = '&download=true&id=%s&filename=%s' % (self.id, filename)
        url = content + download

        return {'type': 'ir.actions.act_url', 'url': url,}