import io
import xlrd
import json
import pytz
import base64
import openpyxl 
import itertools
import calendar
from collections import OrderedDict
from xlwt import easyxf
from odoo.http import request
from odoo.tools.misc import xlwt
from odoo.tools import date_utils
from odoo import api, fields, models, _
from odoo.exceptions import Warning as UserError
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
from odoo.tools.misc import DEFAULT_SERVER_DATETIME_FORMAT
from openpyxl.utils import get_column_letter, column_index_from_string
from pytz import timezone
try:
    from odoo.tools.misc import xlsxwriter
except ImportError:
    import xlsxwriter


class wizard_report_cogs(models.TransientModel):
    _name = 'wizard.report.cogs'
    _description = "Report COGS"

    date_start = fields.Date("Date Start")
    date_end = fields.Date(string='End Date')
    fleg_sum_so = fields.Boolean("Group SO Origin and SO Marketing", default=True)
    data_file = fields.Binary('File')
    name = fields.Char('File Name')

    @api.onchange('date_start')
    def _onchange_date_start(self):
        if self.date_start:
            self.date_end = self.date_start + relativedelta(months=1) - relativedelta(days=1)

    def eksport_excel(self):
        book = xlwt.Workbook(encoding="utf-8")
        title = self._description

        sheet = book.add_sheet(title, cell_overwrite_ok=True)
        sheet.normal_magn = 80
        # sheet.show_grid = False

        style_header2 = xlwt.easyxf('font: name Calibri, height 210, bold 1; \
            pattern: pattern solid, fore_colour gray25;borders: left thin, \
            right thin, top thin, bottom thin;align: vert centre, horz center', num_format_str='#,##0.0000')
        style_no_bold = xlwt.easyxf('font: name Calibri; align: vert centre, horz center')
        style_no_bold.num_format_str = '#,##0.0000'

        now = datetime.now()
        judul = "Report COGS"
        header = ['PERIOD','CATEG PERIOD','INTERNAL REFERENCE','PRODUCT','SALES QTY', 'SALE VALUE', 'COGS','RM','DL','DEPRE','FOH','COGS/unit','RM/unit','DL/unit','DEPRE/unit','FOH/unit']
        colh = -1
        for x in header:
            colh += 1
            style_header2.alignment.wrap = 1
            sheet.write(5, colh, x, style_header2)

        date_from = self.date_start or "0001-01-01"
        date_to = self.date_end or fields.Date.context_today(self)
        where_date_from = "so.date_order >= '%s 00:00:00'" % (date_from)
        where_date_to = "so.date_order <= '%s 23:59:59'" % (date_to)

        setting = self.env['res.company']._company_default_get('mrp.cost.settings')
        product_obj = self.env['product.product']
        so_obj = self.env['sale.order']
        sol_obj = self.env['sale.order.line']
        picking_obj = self.env['stock.picking']
        ml_obj = self.env['stock.move.line']
        data_dict = {}
        data = []
        
        sql_query = """
            SELECT
                so.sales_type AS sales_type,
				move_line.product_id,
                move_line.product_uom_id,
                move_line.location_id, 
                move_line.location_dest_id, 
                sum(move_line.qty_done) AS qty_done
            FROM stock_move_line move_line
                JOIN stock_picking AS picking on picking.id = move_line.picking_id
                JOIN sale_order AS so on so.id = picking.sale_id
            WHERE move_line.state = 'done' and %s and %s
            GROUP BY so.sales_type, move_line.product_id, move_line.product_uom_id, move_line.location_id, move_line.location_dest_id
            ORDER BY move_line.product_id
        """% (where_date_from, where_date_to)
        self._cr.execute(sql_query)
        result = self._cr.dictfetchall()
        for l in result:
            product = self.env['product.product'].browse(l['product_id'])
            product_uom = self.env['uom.uom'].browse(l['product_uom_id'])
            location = self.env['stock.location'].browse(l['location_id'])
            location_dest = self.env['stock.location'].browse(l['location_dest_id'])
            sales_type = l['sales_type']
            
            if product_uom.id != product.uom_id.id:
                qty_done = product_uom._compute_quantity(l['qty_done'], product.uom_id)
            else:
                qty_done = l['qty_done']

            if location.usage == 'customer':
                qty_done = -qty_done
            
            if product.id not in data_dict:
                data_dict[product.id] = [qty_done] 
            else:
                data_dict[product.id].append(qty_done)
            
        for w, y in data_dict.items():
            product = product_obj.browse(w)
            qty = sum(y)
            unit_price = product.lst_price
            sql_so = """
                SELECT
                    sol.id
                FROM sale_order_line sol
                    JOIN sale_order AS so on so.id = sol.order_id
                WHERE so.state in ('sale','done') and %s and %s and sol.product_id = %s
            """% (where_date_from, where_date_to, w)
            self._cr.execute(sql_so)
            result_so = self._cr.dictfetchall()
            price_subtotal = 0
            for p in result_so:
                price_subtotal += self.env['sale.order.line'].browse(p['id']).price_subtotal
            # price_subtotal = qty * unit_price
            cost_line_id = product.cost_line_id
            cost_id = product.cost_line_id.cost_id
            rm_persen = dl_persen = depre_persen = foh_persen = hasil_rm = hasil_dl = hasil_depre = hasil_foh = cogs_unit = rm_unit = dl_unit = depre_unit = foh_unit = 0
            period = ''
            cogs_unit = product.standard_price
            pmac = cogs_unit * qty
            if cost_line_id and cost_id.state == 'done':
                mrp_selection_month = dict(cost_id.fields_get(['select_month'])['select_month']['selection'])
                month = [v for k,v in mrp_selection_month.items() if k == cost_id.select_month ][0]
                mrp_selection_year = dict(cost_id.fields_get(['select_year'])['select_year']['selection'])
                year = [v for k,v in mrp_selection_year.items() if k == cost_id.select_year ][0]
                period = f'{month} - {year}'
                # amount_total = cost_id.amount_total
                mrp_cost = self.env['mrp.cost'].search([('select_month','=',cost_id.select_month),('select_year','=',cost_id.select_year),('state','=','done')])
                amount_total = sum([x.amount_total for x in mrp_cost])
                amount_rm = sum([x.price_unit for x in self.env['mrp.cost.line'].search([('cost_id','in',mrp_cost.ids),('product_id','in',setting.product_rm_ids.ids)])])
                amount_dl = sum([x.price_unit for x in self.env['mrp.cost.line'].search([('cost_id','in',mrp_cost.ids),('product_id','in',setting.product_dl_ids.ids)])])
                amount_depre = sum([x.price_unit for x in self.env['mrp.cost.line'].search([('cost_id','in',mrp_cost.ids),('product_id','in',setting.product_depre_ids.ids)])])
                amount_foh = sum([x.price_unit for x in self.env['mrp.cost.line'].search([('cost_id','in',mrp_cost.ids),('product_id','in',setting.product_foh_ids.ids)])])
                if amount_rm > 0:
                    rm_persen = amount_rm / amount_total
                    hasil_rm = rm_persen * pmac
                if amount_dl > 0:
                    dl_persen = amount_dl / amount_total
                    hasil_dl = dl_persen * pmac
                if amount_depre > 0:
                    depre_persen = amount_depre / amount_total
                    hasil_depre = depre_persen * pmac
                if amount_foh > 0:
                    foh_persen = amount_foh / amount_total
                    hasil_foh = foh_persen * pmac
                rm_unit = hasil_rm / qty
                dl_unit = hasil_dl / qty
                depre_unit = hasil_depre / qty
                foh_unit = hasil_foh / qty
            default_code = product.default_code
            product_name = product.name
            data.append({
                'PERIOD'            : period,
                'CATEG PERIOD'      : product.categ_id.name,
                'INTERNAL REFERENCE': default_code,
                'PRODUCT'           : product_name,
                'SALES QTY'         : qty,
                'SALE VALUE'        : price_subtotal,
                'COGS'              : pmac,
                'RM'                : hasil_rm,
                'DL'                : hasil_dl,
                'DEPRE'             : hasil_depre,
                'FOH'               : hasil_foh,
                'COGS/unit'         : cogs_unit,
                'RM/unit'           : rm_unit,
                'DL/unit'           : dl_unit,
                'DEPRE/unit'        : depre_unit,
                'FOH/unit'          : foh_unit,
            })
        no = 5
        for line in data:
            no += 1
            col = -1
            for i in header:
                col += 1
                sheet.write(no, col, line[i], style_no_bold)
        sheet.write(len(data)+6, 3, 'Total', style_header2)
        num = 4
        col = 3
        for x in range(len(header)):
            num += 1
            col += 1
            if num < len(header) + 1:
                cek_col = get_column_letter(num)
                sum_volume = "SUM(%s%d:%s%d)" % (cek_col, 7, cek_col, 6+len(data_dict))
                sheet.write(no+1, col, xlwt.Formula(sum_volume), style_header2)
        col_width = 256 * 25
        try:
            for i in itertools.count():
                sheet.col(i).width = col_width
                sheet.col(0).width = 256 * 20
                sheet.col(1).width = 256 * 30
                sheet.col(2).width = 256 * 30
                sheet.col(3).width = 256 * 50
                sheet.col(4).width = 256 * 30
        except ValueError:
            pass

        file_data = io.BytesIO()
        book.save(file_data)

        self.name = '%s %s.xls' % (judul, now)

        excel_file = base64.encodestring(file_data.getvalue())
        self.data_file = excel_file

        view = self.env.ref('bmo_mrp_cost.wizard_report_cogs_view_form')

        return {
            'view_type': 'form',
            'views': [(view.id, 'form')],
            'view_mode': 'form',
            'res_id': self.id,
            'res_model': 'wizard.report.cogs',
            'type': 'ir.actions.act_window',
            'target': 'new',
        }
